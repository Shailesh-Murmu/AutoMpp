import sys
import os
import json
import logging
import smtplib
import io
import pandas as pd
import html
import datetime
import re
import base64
import requests # Add this import for downloading the default image
import webbrowser # Add this import to open web links

# --- Third-party libraries ---
# Make sure to install them: pip install PyQt5 openpyxl google-api-python-client google-auth-oauthlib google-auth-httplib2 python-dotenv google-generativeai requests
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QTextEdit,
    QFileDialog, QVBoxLayout, QHBoxLayout, QMessageBox, QTabWidget, QComboBox, QGridLayout, QDateEdit,
    QProgressDialog, QSlider, QDesktopWidget
)
from email.mime.text import MIMEText
from PyQt5.QtCore import Qt, QObject, QThread, pyqtSignal, pyqtSlot, QPropertyAnimation, QEasingCurve, QRect, QBuffer
from PyQt5.QtGui import QColor, QIcon, QFont, QPalette, QMovie, QPainter, QBrush, QPixmap
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from googleapiclient.discovery import build
from googleapiclient import errors
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.http import MediaIoBaseDownload
import shutil
from dotenv import load_dotenv
import google.generativeai as genai


# --- Configuration ---
load_dotenv()

CONFIG_FILE = 'task_log.json'
TOKEN_FILE = 'token.json'
CREDENTIALS_FILE = 'credentials.json'
LOGFILE = 'automation.log'
DEFAULT_BG_FILE = 'b.jpg' # Filename for the default background
SMTP_EMAIL = os.environ.get('AUTOMATION_SMTP_EMAIL')
SMTP_PASSWORD = os.environ.get('AUTOMATION_SMTP_PASSWORD')
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SCOPES = ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/forms.body']

logging.basicConfig(filename=LOGFILE, level=logging.INFO, format='%(asctime)s - GUI - %(levelname)s - %(message)s')

# --- Base64 Embedded Assets ---
# A simple loading spinner GIF
DEFAULT_LOADING_SPINNER_BASE64 = "R0lGODlhEAAQAPIAAP///wAAAMLCwkJCQgAAAGJiYoKCgpKSkiH/C05FVFNDQVBFMi4wAwEAAAAh/hpDcmVhdGVkIHdpdGggYWpheGxvYWQuaW5mbwAh+QQJCgAAACwAAAAAEAAQAAADMwi63P4wyklrE2MIOggZnAdOmGYJRbExwroUmcG2LmDEwnHQLVsYOd2mBzkYDAdKa+dIAAAh+QQJCgAAACwAAAAAEAAQAAADNAi63P5OjCEgG4QMu7DmikRxQlFUYDEZIGBMRVsaqHwctXXf7WEYB4Ag1xjihkMZsiUkKhIAIfkECQoAAAAsAAAAABAAEAAAAzYIujIjK8pByJDMlFYvBoVjHA70GU7xSUJhmKtwHPAKzLO9HMaoKwJZ7Rf8AYPDDzKpZBqfvwQAIfkECQoAAAAsAAAAABAAEAAAAzMIumIlK8oyhpHsnFZfhYumCYUhDAQxRIdhHBGqRoKw0R8DYlJd8z0fMDgsGo/IpHI5TAAAIfkECQoAAAAsAAAAABAAEAAAAzIIunInK0rnZBTwGPNMgQwmdsNgXGJUlIWEuR5oWUIpz8pAEAMe6TwfwyYsGo/IpFKSAAAh+QQJCgAAACwAAAAAEAAQAAADMwi6IMKQORfjdOe82p4wGccc4CEuQradylesojEMBgsUc2G7sDX3lQGBMLAJibufbSlKAAAh+QQJCgAAACwAAAAAEAAQAAADMgi63P7wCRHZnFVdmgHu2nFwlWCI3WGc3TSWhUFGxTAUkGCbtgENBMJAEJsxgMLWzpEAACH5BAkKAAAALAAAAAAQABAAAAMyCLrc/jDKSatlQtScKdceCAjDII7HcQ4EMTCpyrCuUBjCYRgHVtqlAiB1YhiCnlsRkAAAOwAAAAAAAAAAAA=="

# --- Global Functions ---
def download_default_background():
    """Downloads a default background image from Unsplash if it doesn't exist."""
    if not os.path.exists(DEFAULT_BG_FILE):
        try:
            # A high-quality, royalty-free image from Unsplash by Fakurian Design
            url = "https://images.unsplash.com/photo-1550745165-9bc0b252726a?q=80&w=1920"
            response = requests.get(url, stream=True)
            response.raise_for_status()
            with open(DEFAULT_BG_FILE, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            logging.info(f"Downloaded default background to {DEFAULT_BG_FILE}")
        except Exception as e:
            logging.error(f"Could not download default background image: {e}")

def get_google_id_from_url(url):
    """Extracts the ID from a Google Drive/Sheets/Forms URL using regex."""
    patterns = [
        r'/d/([a-zA-Z0-9-_]+)',
        r'/folders/([a-zA-Z0-9-_]+)',
        r'id=([a-zA-Z0-9-_]+)'
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return url

def get_creds():
    creds = None
    if not os.path.exists(CREDENTIALS_FILE):
        QMessageBox.critical(None, "Error", f"Credentials file ('{CREDENTIALS_FILE}') not found.")
        return None
    if os.path.exists(TOKEN_FILE):
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        except Exception as e:
            logging.warning(f"Failed to load token.json: {e}. Will try to re-authenticate.")
            creds = None
            
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                QMessageBox.critical(None, "Error", f"Could not refresh token. You may need to re-authenticate. Error: {e}")
                creds = None
        if not creds:
            try:
                flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
                creds = flow.run_local_server(port=0)
            except Exception as e:
                QMessageBox.critical(None, "Error", f"Failed to authenticate with Google. Please check your '{CREDENTIALS_FILE}'. Error: {e}")
                return None
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())
    return creds

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f: return json.load(f)
        except Exception:
            shutil.copy(CONFIG_FILE, CONFIG_FILE + ".backup"); return {}
    return {}

def save_config(config):
    with open(CONFIG_FILE, 'w') as f: json.dump(config, f, indent=4)

def format_markdown_to_html(text):
    """Converts a simple markdown-like text to basic HTML for display."""
    text = html.escape(text)
    text = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', text)
    lines = text.split('\n')
    in_list = False
    formatted_lines = []
    for line in lines:
        match = re.match(r'^\s*(\d+)\.\s*(.*)', line)
        if match:
            if not in_list:
                formatted_lines.append('<ol>')
                in_list = True
            formatted_lines.append(f'<li>{match.group(2)}</li>')
        else:
            if in_list:
                formatted_lines.append('</ol>')
                in_list = False
            formatted_lines.append(line)
    if in_list:
        formatted_lines.append('</ol>')
    return '<br>'.join(formatted_lines).replace('<br><ol>', '<ol>').replace('</ol><br>', '</ol>')


# --- Loading Overlay Widget ---
class LoadingOverlay(QWidget):
    """A semi-transparent overlay with a loading spinner."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setVisible(False)
        self.spinner_movie = QMovie(self)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)

        self.spinner_label = QLabel(self)
        self.spinner_label.setMovie(self.spinner_movie)
        
        self.loading_text = QLabel("Diagnosing with AI...", self)
        self.loading_text.setAlignment(Qt.AlignCenter)
        self.loading_text.setStyleSheet("color: white; font-size: 14pt; font-weight: 600;")

        layout.addWidget(self.spinner_label)
        layout.addWidget(self.loading_text)
        self.setLayout(layout)

    def set_spinner(self, gif_path=None):
        """Sets the spinner GIF from a path or falls back to the default."""
        self.spinner_movie.stop()
        
        spinner_data = None
        if gif_path and os.path.exists(gif_path):
            try:
                with open(gif_path, 'rb') as f:
                    spinner_data = f.read()
            except Exception as e:
                logging.error(f"Failed to load custom spinner GIF: {e}")
                spinner_data = None

        if not spinner_data:
            spinner_data = base64.b64decode(DEFAULT_LOADING_SPINNER_BASE64)

        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        buffer.write(spinner_data)
        buffer.seek(0)
        
        self.spinner_movie.setDevice(buffer)
        self.spinner_movie.setFormat(b'gif')
        # We need to manage the buffer's lifetime ourselves
        self.spinner_buffer = buffer


    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setBrush(QBrush(QColor(0, 0, 0, 150)))
        painter.setPen(Qt.NoPen)
        painter.drawRect(self.rect())

    def show(self):
        if self.parent():
            self.setGeometry(self.parent().rect())
        self.spinner_movie.start()
        super().show()

    def hide(self):
        self.spinner_movie.stop()
        super().hide()

# --- Worker Objects for Threading ---

class GeminiWorker(QObject):
    """Worker to call Gemini API in a separate thread."""
    finished = pyqtSignal(str)

    def __init__(self, api_key, prompt):
        super().__init__()
        self.api_key = api_key
        self.prompt = prompt

    @pyqtSlot()
    def process(self):
        """Executes the Gemini API call. Renamed from 'run' for clarity."""
        try:
            genai.configure(api_key=self.api_key)
            model = genai.GenerativeModel('gemini-1.5-flash-latest')
            response = model.generate_content(self.prompt)
            html_response = format_markdown_to_html(response.text.strip())
            self.finished.emit(html_response)
        except Exception as e:
            logging.warning(f"Could not contact Gemini AI: {e}")
            self.finished.emit(html.escape(str(self.prompt.split("Technical Error: ")[-1])))

class EmailWorker(QObject):
    """Worker to process and send emails in a separate thread."""
    finished = pyqtSignal(str, str)
    progress = pyqtSignal(int, int)
    
    def __init__(self, task, smtp_details):
        super().__init__()
        self.task = task
        self.smtp_details = smtp_details
        self.is_running = True

    def send_email_worker(self, recipient, subject, body, cc_recipients_str=None):
        try:
            msg = MIMEText(body)
            msg['Subject'] = subject
            msg['From'] = self.smtp_details['email']
            msg['To'] = recipient
            all_recipients = [recipient]
            if cc_recipients_str:
                cc_list = [email.strip() for email in cc_recipients_str.split(',') if email.strip()]
                if cc_list:
                    msg['Cc'] = ", ".join(cc_list)
                    all_recipients.extend(cc_list)
            with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
                server.login(self.smtp_details['email'], self.smtp_details['password'])
                server.sendmail(self.smtp_details['email'], all_recipients, msg.as_string())
            logging.info(f"Sent email to {recipient}")
            return True
        except Exception as e:
            logging.error(f"Failed to send email to {recipient}: {e}")
            return False

    @pyqtSlot()
    def process(self):
        """Executes the email sending task. Renamed from 'run' for clarity."""
        try:
            df = pd.read_excel(self.task["excel"])
        except FileNotFoundError:
            self.finished.emit('error', f"Excel file not found at: {self.task['excel']}")
            return
        except Exception as e:
            self.finished.emit('error', f"Error reading Excel file. It may be corrupt or unsupported.\n\nDetails: {e}")
            return

        total_rows = len(df)
        success_count, fail_count = 0, 0
        cc_list = self.task.get("cc", "")
        subject = self.task.get("subject", "Automated Message")
        for i, row in df.iterrows():
            if not self.is_running:
                self.finished.emit('error', 'Task cancelled by user.')
                return
            
            recipient = next((str(row[col]).strip() for col in ['Email', 'Email ID'] if col in row and pd.notna(row[col])), None)
            if not recipient:
                fail_count += 1
                continue
            if self.send_email_worker(recipient, subject, self.task["msg"], cc_recipients_str=cc_list):
                success_count += 1
            else:
                fail_count += 1
            self.progress.emit(i + 1, total_rows)
        
        if self.is_running:
            self.finished.emit('success', f"Email process finished.\n\nSuccessfully sent: {success_count}\nFailed: {fail_count}")

    def stop(self): self.is_running = False

class DriveWorker(QObject):
    """Worker to download files from Google Drive in a separate thread."""
    finished = pyqtSignal(str, str)
    progress = pyqtSignal(int, int)

    def __init__(self, creds, folder_id, path):
        super().__init__()
        self.creds, self.folder_id, self.path = creds, folder_id, path
        self.is_running = True
    
    @pyqtSlot()
    def process(self):
        """Executes the drive download task. Renamed from 'run' for clarity."""
        try:
            drive_service = build('drive', 'v3', credentials=self.creds)
            os.makedirs(self.path, exist_ok=True)
            results = drive_service.files().list(q=f"'{self.folder_id}' in parents and trashed=false", fields="files(id, name, mimeType)").execute()
            items = results.get('files', [])
            if not items:
                self.finished.emit('success', "No files found in the specified Google Drive folder.")
                return
            
            total = len(items)
            for i, file in enumerate(items):
                if not self.is_running:
                    self.finished.emit('error', "Download canceled by user.")
                    return
                
                file_id, file_name, mime_type = file.get('id'), file.get('name'), file.get('mimeType')
                unique_file_path = os.path.join(self.path, file_name)
                fh = io.BytesIO()
                if mime_type.startswith('application/vnd.google-apps'):
                    export_map = {'application/vnd.google-apps.document': {'mime': 'application/pdf', 'ext': '.pdf'},'application/vnd.google-apps.spreadsheet': {'mime': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'ext': '.xlsx'},'application/vnd.google-apps.presentation': {'mime': 'application/vnd.openxmlformats-officedocument.presentationml.presentation', 'ext': '.pptx'},}
                    if mime_type in export_map:
                        export_details = export_map[mime_type]; request = drive_service.files().export_media(fileId=file_id, mimeType=export_details['mime'])
                        base_name, _ = os.path.splitext(unique_file_path); unique_file_path = base_name + export_details['ext']
                    else:
                        logging.warning(f"Skipping unsupported Google App file: {file_name}")
                        continue
                else:
                    request = drive_service.files().get_media(fileId=file_id)
                
                downloader = MediaIoBaseDownload(fh, request); done = False
                # IMPROVED CANCELLATION: Check the flag inside the download chunk loop
                while not done:
                    if not self.is_running:
                        self.finished.emit('error', "Download canceled by user.")
                        return
                    status, done = downloader.next_chunk()
                
                with open(unique_file_path, 'wb') as f: f.write(fh.getbuffer())
                self.progress.emit(i + 1, total)
            
            if self.is_running:
                self.finished.emit('success', "Folder downloaded successfully!")

        except errors.HttpError as e:
            self.finished.emit('error', f"API Error: {e}. Check permissions and folder link.")
        except Exception as e:
            self.finished.emit('error', f"Download failed: {e}")
            
    def stop(self): self.is_running = False

def split_and_flatten_docs(series):
    """Helper function to split comma-separated doc links and flatten the list."""
    all_docs = []
    for item in series:
        if pd.notna(item) and str(item).strip():
            # Split comma-separated links and strip whitespace from each
            links = [link.strip() for link in str(item).split(',')]
            all_docs.extend(links)
    return all_docs

class TrackerWorker(QObject):
    """Worker to generate a tracker Excel file in a separate thread."""
    finished = pyqtSignal(str, str)

    def __init__(self, creds, task_details):
        super().__init__()
        self.creds = creds
        self.task = task_details
        self.is_running = True

    @pyqtSlot()
    def process(self):
        """Executes the tracker generation task. Renamed from 'run' for clarity."""
        try:
            sheet_id = get_google_id_from_url(self.task["response_sheet_id"])
            service = build('sheets', 'v4', credentials=self.creds)

            spreadsheet_metadata = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
            sheets = spreadsheet_metadata.get('sheets', [])
            sheet_name = ""
            for s in sheets:
                if s['properties']['title'].startswith("Form Responses"):
                    sheet_name = s['properties']['title']
                    break
            if not sheet_name and sheets:
                sheet_name = sheets[0]['properties']['title']
            
            if not sheet_name:
                self.finished.emit('error', "Could not find a valid response sheet in the Google Sheet file.")
                return

            data = service.spreadsheets().values().get(spreadsheetId=sheet_id, range=sheet_name).execute().get('values', [])
            
            if not data or len(data) < 1:
                self.finished.emit('error', f"No data or header row in response sheet '{sheet_name}'.")
                return
            
            header = data[0]
            padded_rows = [row + [None] * (len(header) - len(row)) for row in data[1:]]
            df_response = pd.DataFrame(padded_rows, columns=header).loc[:,~pd.DataFrame(padded_rows, columns=header).columns.duplicated()]

            required_master_cols = ['Email ID', 'Location', 'SPOC']
            required_response_cols = ['Email', 'Location', 'Upload the Applicable Documents']
            
            df_master = pd.read_excel(self.task["master_excel"])
            if not all(col in df_master.columns for col in required_master_cols):
                self.finished.emit('error', f"Master Excel missing columns: {required_master_cols}"); return
            if not all(col in df_response.columns for col in required_response_cols):
                self.finished.emit('error', f"Response Sheet missing columns: {required_response_cols}"); return

            df_response['Email'] = df_response['Email'].astype(str).str.strip().str.lower()
            df_response['Location'] = df_response['Location'].astype(str).str.strip().str.lower()
            ts_col = 'Timestamp' if 'Timestamp' in df_response.columns else df_response.columns[0]
            
            # UPDATED: Aggregate responses, keeping document links as a list
            agg_response = (df_response.groupby(['Email', 'Location']).agg(
                docs=('Upload the Applicable Documents', split_and_flatten_docs),
                timestamp=(ts_col, 'max')
            ).reset_index())
            
            response_map = {(row['Email'], row['Location']): (row['docs'], row['timestamp']) for _, row in agg_response.iterrows()}
            
            # UPDATED: Build the final data structure for the Excel file, creating multiple rows for multiple documents
            final_rows = []
            s_no = 1
            for _, row in df_master.iterrows():
                if not self.is_running: self.finished.emit('error', "Tracker generation was cancelled."); return
                key = (str(row['Email ID']).strip().lower(), str(row['Location']).strip().lower())
                docs, timestamp = response_map.get(key, ([], ''))

                if docs:
                    first_row_in_group = True
                    for doc_link in docs:
                        final_rows.append({
                            'S.No.': s_no if first_row_in_group else None,
                            'Location': row['Location'] if first_row_in_group else None,
                            'SPOC': row['SPOC'] if first_row_in_group else None,
                            'Email ID': row['Email ID'] if first_row_in_group else None,
                            'Document Uploaded': doc_link,
                            'Uploaded': 'Yes' if first_row_in_group else None,
                            'Uploaded When': timestamp if first_row_in_group else None
                        })
                        first_row_in_group = False
                    s_no += 1
                else:
                    final_rows.append({
                        'S.No.': s_no,
                        'Location': row['Location'],
                        'SPOC': row['SPOC'],
                        'Email ID': row['Email ID'],
                        'Document Uploaded': '',
                        'Uploaded': 'No',
                        'Uploaded When': ''
                    })
                    s_no += 1
            
            tracker_df = pd.DataFrame(final_rows, columns=['S.No.', 'Location', 'SPOC', 'Email ID', 'Document Uploaded', 'Uploaded', 'Uploaded When'])
            output_path = self.task["result_path"]
            tracker_df.to_excel(output_path, index=False)

            # --- UPDATED: Apply advanced formatting and merging with openpyxl ---
            wb = load_workbook(output_path)
            ws = wb.active
            
            for idx, width in enumerate([8, 20, 25, 30, 50, 12, 25], start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width
            
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            center_align = Alignment(vertical="center", horizontal="center", wrap_text=True)

            # Apply alignment to all columns except 'Document Uploaded' (column 5)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.column != 5:
                        cell.alignment = center_align

            # Apply conditional formatting and merging
            cols_to_merge = [1, 2, 3, 4, 6, 7] # S.No, Location, SPOC, Email, Uploaded, Uploaded When
            current_row_idx = 2  # Start from the first data row (Excel rows are 1-based)

            while current_row_idx <= ws.max_row:
                s_no_cell = ws.cell(row=current_row_idx, column=1)
                if s_no_cell.value is not None:
                    # This is the start of a group. Find how many rows it spans.
                    merge_count = 1
                    scan_row_idx = current_row_idx + 1
                    while scan_row_idx <= ws.max_row and ws.cell(row=scan_row_idx, column=1).value is None:
                        merge_count += 1
                        scan_row_idx += 1
                    
                    # Color the 'Uploaded' cell based on its value for the group
                    uploaded_cell = ws.cell(row=current_row_idx, column=6)
                    if uploaded_cell.value == "Yes":
                        uploaded_cell.fill = green_fill
                    elif uploaded_cell.value == "No":
                        uploaded_cell.fill = red_fill

                    # Merge cells if the group has more than one row
                    if merge_count > 1:
                        end_row_idx = current_row_idx + merge_count - 1
                        for col_idx in cols_to_merge:
                            ws.merge_cells(start_row=current_row_idx, start_column=col_idx, end_row=end_row_idx, end_column=col_idx)
                    
                    current_row_idx += merge_count
                else:
                    # Should not be reached, but as a safeguard
                    current_row_idx += 1

            wb.save(output_path)
            
            if self.is_running:
                self.finished.emit('success', f"Tracker file successfully generated at:\n{output_path}")

        except errors.HttpError as e:
            self.finished.emit('error', f"Google Sheets API Error: {e}. Check Response Sheet link and permissions.")
        except FileNotFoundError as e:
            self.finished.emit('error', f"File not found: {e.filename}")
        except Exception as e:
            if self.is_running: self.finished.emit('error', f"Failed to generate tracker: {e}")

    def stop(self): self.is_running = False

class FormWorker(QObject):
    """Worker to update Google Form dropdowns in a separate thread."""
    finished = pyqtSignal(str, str)

    def __init__(self, creds, form_id, excel_path):
        super().__init__()
        self.creds, self.form_id, self.excel_path = creds, form_id, excel_path
        self.is_running = True

    @pyqtSlot()
    def process(self):
        """Executes the form update task. Renamed from 'run' for clarity."""
        try:
            df = pd.read_excel(self.excel_path)
            field_mappings = {'Location': 'Location', 'Email': 'Email ID', 'SPOC Name': 'SPOC'}
            forms_service = build('forms', 'v1', credentials=self.creds)
            form = forms_service.forms().get(formId=self.form_id).execute()
            form_items = form.get('items', [])
            requests, updated_fields = [], []
            for form_title, excel_column in field_mappings.items():
                if not self.is_running: self.finished.emit('error', "Operation cancelled."); return
                if excel_column not in df.columns: logging.warning(f"Column '{excel_column}' not in Excel. Skipping '{form_title}'."); continue
                options_list = df[excel_column].dropna().unique().tolist()
                if not options_list: logging.warning(f"No data in column '{excel_column}'. Skipping '{form_title}'."); continue
                target_item, target_index = None, -1
                for i, item in enumerate(form_items):
                    if item.get('title', '').strip().lower() == form_title.lower():
                        target_item, target_index = item, i; break
                if not target_item: logging.warning(f"Question '{form_title}' not in form. Skipping."); continue
                dropdown_options = [{'value': str(opt)} for opt in options_list]
                new_item_body = {
                    'itemId': target_item.get('itemId'), 'title': target_item.get('title'),
                    'questionItem': {'question': {
                        'questionId': target_item['questionItem']['question']['questionId'],
                        'required': True,
                        'choiceQuestion': {'type': 'DROP_DOWN', 'options': dropdown_options, 'shuffle': False}
                    }}}
                request = {'updateItem': {'item': new_item_body, 'location': {'index': target_index}, 'updateMask': 'questionItem'}}
                requests.append(request)
                updated_fields.append(form_title)
            if not requests: self.finished.emit('error', "No matching questions or data to update."); return
            body = {'requests': requests}
            forms_service.forms().batchUpdate(formId=self.form_id, body=body).execute()
            if self.is_running:
                self.finished.emit('success', f"Successfully updated dropdowns for: {', '.join(updated_fields)}.")
        except errors.HttpError as e: self.finished.emit('error', f"Google Forms API Error: {e}. Check permissions.")
        except FileNotFoundError as e: self.finished.emit('error', f"Master Excel file not found: {e.filename}")
        except Exception as e:
            if self.is_running: self.finished.emit('error', f"An unexpected error occurred: {e}")
    def stop(self): self.is_running = False


class AutomationApp(QWidget):
    STYLESHEET_TEMPLATE = """
        QTabWidget::pane { border: 1px solid #444953; border-top: none; border-radius: 0px 0px 8px 8px; background-color: rgba(53, 58, 69, 0.95); padding: 15px; }
        QTabBar::tab { background: rgba(53, 58, 69, 0.9); color: #d0d0d0; border: 1px solid #444953; border-bottom: none; padding: 10px 25px; font-weight: 600; border-top-left-radius: 8px; border-top-right-radius: 8px; margin-right: 2px; }
        QTabBar::tab:selected { background: #4a8df8; color: #ffffff; border-color: #4a8df8; }
        QTabBar::tab:!selected:hover { background: #4f5563; }
        QPushButton { background-color: #4f5563; color: #ffffff; border: none; padding: 10px 20px; border-radius: 6px; font-weight: 600; }
        QPushButton:hover { background-color: #5a6170; }
        QPushButton:pressed { background-color: #454b57; }
        QPushButton#PrimaryButton { background-color: #4a8df8; }
        QPushButton#PrimaryButton:hover { background-color: #5c9bff; }
        QPushButton#PrimaryButton:pressed { background-color: #3a7de0; }
        QPushButton#DeleteButton { background-color: #e74c3c; }
        QPushButton#DeleteButton:hover { background-color: #ff6b5b; }
        QPushButton#DeleteButton:pressed { background-color: #c0392b; }
        QLineEdit, QTextEdit, QComboBox, QDateEdit { background-color: #2c313c; color: #f0f0f0; border: 1px solid #444953; border-radius: 6px; padding: 8px; }
        QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QDateEdit:focus { border: 1px solid #4a8df8; background-color: #353a45; }
        QTextEdit { font-family: "Consolas", "Courier New", monospace; }
        QComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: top right; width: 25px; border-left-width: 1px; border-left-color: #444953; border-left-style: solid; border-top-right-radius: 6px; border-bottom-right-radius: 6px; }
        QComboBox::down-arrow { image: url(data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxNiIgaGVpZ2h0PSIxNiIgZmlsbD0iI2YwZjBmMCIgY2xhc3M9ImJpIGJpLWNoZXZyb24tZG93biIgdmlld0JveD0iMCAwIDE2IDE2Ij4KICA8cGF0aCBmaWxsLXJ1bGU9ImV2ZW5vZGQiIGQ9Ik0xLjYzNiA0LjEzNmExIDEgMCAwIDEgMS40MTQgMGw0Ljk1IDQuOTUgNC45NS00Ljk1YTEgMSAwIDEgMSAxLjQxNCAxLjQxNGwtNS42NTYgNS42NTdhMSAxIDAgMCAxLTEuNDE0IDBsLTUuNjU3LTUuNjU3YTEgMSAwIDAgMSAwLTEuNDE0eiIvPgo8L3N2Zz4=); width: 14px; height: 14px; }
        QComboBox QAbstractItemView { background-color: #353a45; border: 1px solid #4a8df8; selection-background-color: #4a8df8; color: #f0f0f0; padding: 5px; }
        QDateEdit::down-arrow { image: url(data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxNiIgaGVpZ2h0PSIxNiIgZmlsbD0iI2YwZjBmMCIgY2xhc3M9ImJpIGJpLWNhbGVuZGFyLWV2ZW50IiB2aWV3Qm94PSIwIDAgMTYgMTYiPgogIDxwYXRoIGQ9Ik0xMSAxYS41LjUgMCAwIDEgLjUuNW0wIDJWLjVBLjUuNSAwIDAgMSAxMiAwYy4yNzYgMCAuNS4yMjQgLjUuNXMyLjUgMCAyLjUgMCAwIDAgLjUgLjVWMy41aC00em0tMy4weMTAgMCAxLjE0Ni4wNTQgMS4xNDYuMDk0djEuMTQ2aC0xLjE0NlpNMSA0LjV2OS41YzAgLjI3Ni4yMjQuNS41LjVoMTNjLjI3NiAwIC41LS4yMjQuNS0uNVY0LjV6TTIgMS41YS41LjUgMCAwIDAtLjUuNVYzLjVoM2wtMy0zaDBaTTQuNSAyaC0zVjEuNWgwVjJ6TTQgMHYuNWg0VjBoLTFabTAgMHYuNWg0VjBoLTFaIi8+CiAgPHBhdGggZD0iTTQgNi41YS41LjUgMCAwIDEgLjUuNWg1YS41LjUgMCAwIDEgMCAxSDQuNWEuNS41IDAgMCAxLS41LS41Ii8+Cjwvc3ZnPg==); }
        QCalendarWidget QWidget { background-color: #2c313c; }
        QCalendarWidget QToolButton { color: white; }
        QLabel { color: #d0d0d0; padding: 5px 0; background-color: transparent; }
        QMessageBox { background-color: #353a45; }
        QMessageBox QLabel { color: #f0f0f0; font-size: 11pt; }
        QScrollBar:vertical { border: none; background: #2c313c; width: 12px; margin: 0px 0px 0px 0px; }
        QScrollBar::handle:vertical { background: #4f5563; min-height: 20px; border-radius: 6px; }
        QScrollBar::handle:vertical:hover { background: #5a6170; }
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
        QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: none; }
        QSlider::groove:horizontal { border: 1px solid #444953; height: 8px; background: #2c313c; margin: 2px 0; border-radius: 4px; }
        QSlider::handle:horizontal { background: #4a8df8; border: 1px solid #4a8df8; width: 16px; margin: -4px 0; border-radius: 8px; }
        QSlider::handle:horizontal:hover { background: #5c9bff; border-color: #5c9bff; }
    """
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Automation System Pro"); self.setMinimumSize(950, 700)
        self.config = load_config(); self.thread = self.worker = None; 
        self.background_pixmap = QPixmap()
        self.background_opacity = 0.3 # Default overlay opacity
        
        self.setStyleSheet(self.STYLESHEET_TEMPLATE)
        self.setWindowOpacity(0.0)
        
        # CRITICAL FIX: Initialize the overlay BEFORE the UI that uses it.
        self.loading_overlay = LoadingOverlay(self) 
        self.init_ui()

    def paintEvent(self, event):
        """Custom paint event to draw the background image and overlay."""
        painter = QPainter(self)
        if not self.background_pixmap.isNull():
            # Draw the background pixmap, scaled to fit the window
            painter.drawPixmap(self.rect(), self.background_pixmap)
        
        # Draw a semi-transparent overlay on top
        painter.setBrush(QColor(44, 49, 60, int(255 * self.background_opacity))) # #2c313c with opacity
        painter.setPen(Qt.NoPen)
        painter.drawRect(self.rect())

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.loading_overlay.setGeometry(self.rect())

    def showEvent(self, event):
        super().showEvent(event)
        if not hasattr(self, '_faded_in'):
            self._faded_in = True
            self.fade_in()

    def fade_in(self):
        self.animation = QPropertyAnimation(self, b"windowOpacity")
        self.animation.setDuration(400)
        self.animation.setStartValue(0.0)
        self.animation.setEndValue(1.0)
        self.animation.setEasingCurve(QEasingCurve.InOutCubic)
        self.animation.start(QPropertyAnimation.DeleteWhenStopped)

    def show_error(self, msg):
        if self.thread and self.thread.isRunning():
            logging.warning("A task is running. Displaying raw error instead of starting AI diagnosis.")
            self.display_error_messagebox(f"<i>A task is already in progress.</i><br><br><b>Original error:</b><br>{html.escape(str(msg))}")
            return
            
        gemini_key = self.config.get('settings', {}).get('gemini_api_key', GEMINI_API_KEY)
        if not gemini_key:
            self.display_error_messagebox(str(msg))
        else:
            prompt = f"""You are an expert technical assistant for a desktop automation app. Analyze the following technical error message and explain it in simple, human-readable terms. Respond ONLY in the format: **Tab:** [Tab Name] **Problem:** [Simple explanation] **Likely Cause:** [Probable reason] **Solution:** [Clear steps to fix] **Limitation:** [App limitation or None]. Technical Error: "{msg}" """
            
            self.loading_overlay.show()
            
            self.thread = QThread()
            self.worker = GeminiWorker(gemini_key, prompt)
            self.worker.moveToThread(self.thread)
            
            self.thread.started.connect(self.worker.process)
            self.worker.finished.connect(self.on_gemini_finished)
            self.worker.finished.connect(self.thread.quit)

            # Cleanup is tied to the thread's lifecycle to prevent race conditions
            self.thread.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            self.thread.finished.connect(self.cleanup_thread) 
            
            self.thread.start()

    def on_gemini_finished(self, html_response):
        """Slot to handle the result from the GeminiWorker."""
        self.loading_overlay.hide()
        self.display_error_messagebox(html_response)

    def display_error_messagebox(self, content):
        """Displays the final error message box with formatted content."""
        logging.error(content)
        
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Critical)
        msg_box.setText("An Error Occurred")
        msg_box.setTextFormat(Qt.RichText)
        msg_box.setInformativeText(content)
        msg_box.setWindowTitle("Error")
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec_()

    def save_and_reload(self):
        save_config(self.config)
        self.refresh_dropdowns()

    def init_ui(self):
        self.tabs = QTabWidget(self)
        self.setup_dashboard_tab()
        self.setup_email_tab()
        self.setup_drive_tab()
        self.setup_tracker_tab()
        self.setup_form_updater_tab()
        self.setup_reminder_tab()
        self.setup_log_tab()
        self.setup_settings_tab()
        
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(self.tabs)
        main_layout.setContentsMargins(10, 10, 10, 10)
        self.setLayout(main_layout)
        
        self.refresh_dropdowns()
        self.load_settings()

    def cancel_task(self):
        """Signals the worker to stop and provides immediate user feedback."""
        if self.worker and hasattr(self.worker, 'stop'):
            logging.info("User cancelled task. Signalling worker to stop.")
            self.worker.stop()
            if self.progress:
                # IMPROVED FEEDBACK: Give clear status during cancellation.
                self.progress.setLabelText("Stopping task, please wait...")
                self.progress.setCancelButton(None) # Prevent multiple cancel clicks

    def closeEvent(self, event):
        if self.thread and self.thread.isRunning():
            reply = QMessageBox.question(self, 'Task in Progress', 
                                         "A task is running. Are you sure you want to quit?", 
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                logging.info("User chose to quit. Attempting to stop worker thread.")
                self.cancel_task()
                self.thread.quit()
                if not self.thread.wait(5000): # Wait for 5 seconds
                    logging.warning("Thread did not terminate gracefully. Forcing termination.")
                    self.thread.terminate()
                    self.thread.wait() # Wait for termination to complete
                logging.info("Thread finished. Closing application.")
                event.accept()
            else:
                logging.info("User cancelled application exit.")
                event.ignore()
        else:
            event.accept()

    def setup_dashboard_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)
        
        refresh_button = QPushButton("🔄 Refresh Dashboard")
        refresh_button.clicked.connect(self.refresh_dashboard_display)
        
        self.dashboard_display = QTextEdit()
        self.dashboard_display.setReadOnly(True)
        self.dashboard_display.setStyleSheet("QTextEdit { background-color: #2c313c; border: 1px solid #444953; font-size: 11pt; }")
        
        layout.addWidget(refresh_button)
        layout.addWidget(self.dashboard_display)
        self.tabs.addTab(tab, "📊 Dashboard")

    def refresh_dashboard_display(self):
        self.config = load_config()
        html_content = """
        <html><head><style>
            body { color: #f0f0f0; font-family: 'Segoe UI'; }
            h2 { color: #4a8df8; border-bottom: 1px solid #444953; padding-bottom: 5px; }
            ul { list-style-type: none; padding-left: 0; }
            li { background-color: #353a45; margin-bottom: 8px; padding: 10px; border-radius: 6px; border-left: 4px solid #4a8df8; }
            b { color: #ffffff; }
            i { color: #a0a0a0; }
        </style></head><body>
        """
        task_types = {
            "emails": "📧 Email Tasks", "drive_tasks": "📁 Drive Tasks",
            "track_tasks": "📈 Tracker Tasks", "form_updater_tasks": "📝 Form Updater Tasks",
            "reminders": "🔔 Reminder Tasks"
        }
        for key, title in task_types.items():
            html_content += f"<h2>{title}</h2>"
            tasks = self.config.get(key, [])
            if not tasks:
                html_content += "<p><i>No tasks have been configured for this category.</i></p>"
                continue
            html_content += "<ul>"
            for task in tasks:
                html_content += f"<li><b>{html.escape(task.get('title', 'Untitled'))}</b></li>"
            html_content += "</ul>"
        html_content += "</body></html>"
        self.dashboard_display.setHtml(html_content)

    def setup_log_tab(self):
        tab = QWidget(); layout = QVBoxLayout(tab); layout.setSpacing(10)
        button_layout = QHBoxLayout(); 
        refresh_button = QPushButton("🔄 Refresh Log"); refresh_button.clicked.connect(self.load_log_file); 
        clear_button = QPushButton("🗑️ Clear Log"); clear_button.clicked.connect(self.clear_log_file);
        clear_button.setObjectName("DeleteButton")
        
        button_layout.addWidget(refresh_button); button_layout.addWidget(clear_button); 
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        self.log_display = QTextEdit(); self.log_display.setReadOnly(True); layout.addWidget(self.log_display)
        tab.setLayout(layout); self.tabs.addTab(tab, "📝 Activity Log"); self.load_log_file()
        
    def load_log_file(self):
        if not os.path.exists(LOGFILE): self.log_display.setText("Log file not found."); return
        try:
            with open(LOGFILE, 'r') as f: log_content = f.readlines()
            html_content = ""
            for line in reversed(log_content):
                safe_line = html.escape(line).strip()
                if not safe_line: continue
                
                color = "#d0d0d0"
                if "ERROR" in line or "CRITICAL" in line: color = "#e74c3c"
                elif "WARNING" in line: color = "#f39c12"
                elif "INFO" in line: color = "#2ecc71"
                
                html_content += f'<p style="color:{color}; margin: 2px 0;">{safe_line.replace(os.linesep, "<br>")}</p>'
                
            self.log_display.setHtml(f"<html><body style='font-family: Consolas, monospace; font-size: 9pt;'>{html_content}</body></html>")
            self.log_display.verticalScrollBar().setValue(0)
        except Exception as e: self.log_display.setText(f"Error reading log file: {e}")
        
    def clear_log_file(self):
        reply = QMessageBox.question(self, 'Confirm Clear', "Are you sure you want to permanently clear the log file?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                with open(LOGFILE, 'w') as f: f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - INFO - Log file cleared by user.\n")
                self.load_log_file()
            except Exception as e: self.show_error(f"Could not clear log file: {e}")

    def open_url(self, url):
        """Opens a URL in the user's default web browser."""
        webbrowser.open(url)

    def setup_settings_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        grid = QGridLayout()
        grid.setSpacing(15)
        
        # --- Credentials ---
        grid.addWidget(QLabel("<b>Credentials</b>"), 0, 0, 1, 2)
        
        # SMTP Email
        self.smtp_email_input = QLineEdit()
        grid.addWidget(QLabel("SMTP Email:"), 1, 0); grid.addWidget(self.smtp_email_input, 1, 1)
        
        # SMTP Password
        self.smtp_password_input = QLineEdit()
        self.smtp_password_input.setEchoMode(QLineEdit.Password)
        smtp_pass_layout = QHBoxLayout()
        smtp_pass_layout.addWidget(self.smtp_password_input)
        get_app_pass_btn = QPushButton("Get App Password")
        get_app_pass_btn.clicked.connect(lambda: self.open_url('https://myaccount.google.com/apppasswords'))
        smtp_pass_layout.addWidget(get_app_pass_btn)
        grid.addWidget(QLabel("SMTP App Password:"), 2, 0); grid.addLayout(smtp_pass_layout, 2, 1)
        
        # Gemini API Key
        self.gemini_api_key_input = QLineEdit()
        gemini_key_layout = QHBoxLayout()
        gemini_key_layout.addWidget(self.gemini_api_key_input)
        get_gemini_key_btn = QPushButton("Get API Key")
        get_gemini_key_btn.clicked.connect(lambda: self.open_url('https://aistudio.google.com/app/apikey'))
        gemini_key_layout.addWidget(get_gemini_key_btn)
        grid.addWidget(QLabel("Gemini API Key:"), 3, 0); grid.addLayout(gemini_key_layout, 3, 1)
        
        # --- Aesthetics ---
        grid.addWidget(QLabel("<b>Appearance</b>"), 4, 0, 1, 2)
        
        # Background Image
        self.bg_image_path_input = QLineEdit(); self.bg_image_path_input.setReadOnly(True)
        browse_bg_btn = QPushButton("Browse...")
        browse_bg_btn.clicked.connect(self.browse_for_background)
        bg_layout = QHBoxLayout(); bg_layout.addWidget(self.bg_image_path_input); bg_layout.addWidget(browse_bg_btn)
        grid.addWidget(QLabel("Background Image:"), 5, 0); grid.addLayout(bg_layout, 5, 1)
        
        # Background Opacity
        self.bg_opacity_slider = QSlider(Qt.Horizontal)
        self.bg_opacity_slider.setRange(0, 100) # 0% to 100%
        self.bg_opacity_slider.valueChanged.connect(self.update_opacity_label)
        self.bg_opacity_label = QLabel("50%")
        opacity_layout = QHBoxLayout(); opacity_layout.addWidget(self.bg_opacity_slider); opacity_layout.addWidget(self.bg_opacity_label)
        grid.addWidget(QLabel("Background Opacity:"), 6, 0); grid.addLayout(opacity_layout, 6, 1)
        
        # Loading Spinner
        self.spinner_path_input = QLineEdit(); self.spinner_path_input.setReadOnly(True)
        browse_spinner_btn = QPushButton("Browse...")
        browse_spinner_btn.clicked.connect(self.browse_for_spinner)
        spinner_layout = QHBoxLayout(); spinner_layout.addWidget(self.spinner_path_input); spinner_layout.addWidget(browse_spinner_btn)
        grid.addWidget(QLabel("Loading Spinner (GIF):"), 7, 0); grid.addLayout(spinner_layout, 7, 1)
        
        grid.setColumnStretch(1, 1)
        
        save_btn = QPushButton("Save Settings")
        save_btn.setObjectName("PrimaryButton")
        save_btn.clicked.connect(self.save_settings)
        
        layout.addLayout(grid)
        layout.addWidget(save_btn, 0, Qt.AlignTop)
        layout.addStretch()
        self.tabs.addTab(tab, "⚙️ Settings")

    def load_settings(self):
        settings = self.config.get("settings", {})
        self.smtp_email_input.setText(settings.get("smtp_email", ""))
        self.smtp_password_input.setText(settings.get("smtp_password", ""))
        self.gemini_api_key_input.setText(settings.get("gemini_api_key", ""))
        
        # Load aesthetic settings
        self.bg_image_path_input.setText(settings.get("background_image_path", DEFAULT_BG_FILE))
        self.spinner_path_input.setText(settings.get("spinner_gif_path", "Default"))
        opacity_value = settings.get("background_opacity", 30) # Default to 30%
        self.bg_opacity_slider.setValue(opacity_value)
        self.update_opacity_label(opacity_value)
        
        self.apply_aesthetic_settings()

    def save_settings(self):
        self.config.setdefault("settings", {})
        self.config["settings"]["smtp_email"] = self.smtp_email_input.text().strip()
        self.config["settings"]["smtp_password"] = self.smtp_password_input.text().strip()
        self.config["settings"]["gemini_api_key"] = self.gemini_api_key_input.text().strip()
        
        # Save aesthetic settings
        self.config["settings"]["background_image_path"] = self.bg_image_path_input.text()
        self.config["settings"]["background_opacity"] = self.bg_opacity_slider.value()
        self.config["settings"]["spinner_gif_path"] = self.spinner_path_input.text()
        
        save_config(self.config)
        self.apply_aesthetic_settings() # Apply them immediately
        QMessageBox.information(self, "Success", "Settings saved successfully.")

    def update_opacity_label(self, value):
        self.bg_opacity_label.setText(f"{value}%")

    def browse_for_background(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Background Image", "", "Image Files (*.png *.jpg *.jpeg)")
        if path:
            self.bg_image_path_input.setText(path)

    def browse_for_spinner(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Loading Spinner", "", "GIF Files (*.gif)")
        if path:
            self.spinner_path_input.setText(path)

    def apply_aesthetic_settings(self):
        """Loads and applies the visual settings from the config."""
        settings = self.config.get("settings", {})
        
        # Background Image
        bg_path = settings.get("background_image_path", DEFAULT_BG_FILE)
        if not os.path.exists(bg_path):
            bg_path = DEFAULT_BG_FILE # Fallback
        self.background_pixmap.load(bg_path)
        
        # Opacity
        opacity_percent = settings.get("background_opacity", 30)
        self.background_opacity = opacity_percent / 100.0
        
        # Spinner
        spinner_path = settings.get("spinner_gif_path")
        self.loading_overlay.set_spinner(spinner_path)
        
        self.update() # Trigger a repaint

    def setup_email_tab(self):
        tab = QWidget(); layout = QVBoxLayout(tab); grid = QGridLayout()
        grid.setSpacing(10)
        grid.addWidget(QLabel("Task Title:"), 0, 0); self.email_task_title = QLineEdit(); grid.addWidget(self.email_task_title, 0, 1, 1, 2)
        grid.addWidget(QLabel("Subject:"), 1, 0); self.email_subject = QLineEdit(); grid.addWidget(self.email_subject, 1, 1, 1, 2)
        grid.addWidget(QLabel("Excel File With Emails:"), 2, 0); self.email_excel_path = QLineEdit(); grid.addWidget(self.email_excel_path, 2, 1)
        browse_button = QPushButton("Browse"); browse_button.clicked.connect(lambda: self.email_excel_path.setText(QFileDialog.getOpenFileName(self, "Select Excel file", "", "Excel Files (*.xlsx *.xls)")[0])); grid.addWidget(browse_button, 2, 2)
        grid.addWidget(QLabel("CC (comma-separated):"), 3, 0); self.email_cc = QLineEdit(); self.email_cc.setPlaceholderText("email1@example.com, email2@example.com"); grid.addWidget(self.email_cc, 3, 1, 1, 2)
        grid.addWidget(QLabel("Custom Message:"), 4, 0, Qt.AlignTop); self.email_message = QTextEdit(); self.email_message.setMinimumHeight(100); grid.addWidget(self.email_message, 4, 1, 1, 2)
        grid.addWidget(QLabel("Send On (yyyy-mm-dd):"), 5, 0); self.email_schedule_date = QDateEdit(); self.email_schedule_date.setCalendarPopup(True); self.email_schedule_date.setDisplayFormat("yyyy-MM-dd"); self.email_schedule_date.setDate(datetime.datetime.today()); grid.addWidget(self.email_schedule_date, 5, 1)
        grid.setColumnStretch(1, 1)
        layout.addLayout(grid)
        buttons_layout = QHBoxLayout(); save_btn = QPushButton("Save/Update Task"); save_btn.clicked.connect(self.save_email_task); delete_btn = QPushButton("Delete Task"); delete_btn.setObjectName("DeleteButton"); delete_btn.clicked.connect(self.delete_email_task); buttons_layout.addWidget(save_btn); buttons_layout.addWidget(delete_btn); layout.addLayout(buttons_layout)
        self.email_task_combo = QComboBox(); self.email_task_combo.currentIndexChanged.connect(self.load_selected_email_task); layout.addWidget(QLabel("Select Existing Task to Edit or Delete:")); layout.addWidget(self.email_task_combo)
        self.send_emails_button = QPushButton("📤 Send Emails Now"); self.send_emails_button.setObjectName("PrimaryButton"); self.send_emails_button.clicked.connect(self.send_emails); layout.addWidget(self.send_emails_button)
        self.tabs.addTab(tab, "📧 Email Task")
        
    def reload_email_combo(self):
        current_text = self.email_task_combo.currentText()
        self.email_task_combo.blockSignals(True)
        self.email_task_combo.clear(); self.config.setdefault("emails", []); self.email_task_combo.addItem("--- Create New Task ---"); self.email_task_combo.addItems([t["title"] for t in self.config["emails"]])
        index = self.email_task_combo.findText(current_text)
        if index != -1: self.email_task_combo.setCurrentIndex(index)
        self.email_task_combo.blockSignals(False)
        
    def save_email_task(self):
        title, subject, excel_path, cc, msg, date = self.email_task_title.text().strip(), self.email_subject.text().strip(), self.email_excel_path.text().strip(), self.email_cc.text().strip(), self.email_message.toPlainText().strip(), self.email_schedule_date.date().toString("yyyy-MM-dd")
        if not all([title, subject, excel_path, msg]): self.show_error("Title, Subject, Excel File, and a Custom Message are required."); return
        task = {"title": title, "subject": subject, "excel": excel_path, "cc": cc, "msg": msg, "date": date, "type": "email"}
        emails = self.config.setdefault("emails", []); self.config["emails"] = [t for t in emails if t["title"] != title]; self.config["emails"].append(task)
        self.save_and_reload(); QMessageBox.information(self, "Saved", "Email task saved successfully.")
        self.email_task_combo.setCurrentText(title)

    def delete_email_task(self):
        title = self.email_task_combo.currentText()
        if not title or title == "--- Create New Task ---": self.show_error("Select a task to delete."); return
        reply = QMessageBox.question(self, 'Confirm Delete', f"Are you sure you want to delete the task '{title}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No: return
        emails = self.config.get("emails", []); self.config["emails"] = [t for t in emails if t["title"] != title]
        self.save_and_reload(); 
        self.load_selected_email_task()
        QMessageBox.information(self, "Deleted", "Email task deleted.")
        
    def load_selected_email_task(self):
        title = self.email_task_combo.currentText()
        obj = next((t for t in self.config.get("emails", []) if t["title"] == title), None)
        if obj:
            self.email_task_title.setText(obj.get("title", "")); self.email_subject.setText(obj.get("subject", "")); self.email_excel_path.setText(obj.get("excel", "")); self.email_cc.setText(obj.get("cc", "")); self.email_message.setPlainText(obj.get("msg", ""))
            self.email_schedule_date.setDate(datetime.datetime.strptime(obj.get("date", datetime.datetime.today().strftime("%Y-%m-%d")), "%Y-%m-%d").date())
        else:
            self.email_task_title.clear(); self.email_subject.clear(); self.email_excel_path.clear(); self.email_cc.clear(); self.email_message.clear(); self.email_schedule_date.setDate(datetime.datetime.today())
            
    def send_emails(self):
        if self.thread and self.thread.isRunning(): self.show_error("A task is already running."); return
        title = self.email_task_title.text().strip()
        task = next((t for t in self.config.get("emails", []) if t["title"] == title), None)
        if not task: self.show_error("Please select or save a task first."); return
        
        settings = self.config.get("settings", {})
        smtp_details = {
            'email': settings.get('smtp_email', SMTP_EMAIL),
            'password': settings.get('smtp_password', SMTP_PASSWORD)
        }
        if not smtp_details['email'] or not smtp_details['password']:
            self.show_error("SMTP credentials not configured in Settings tab."); return

        self.progress = QProgressDialog("Processing emails...", "Cancel", 0, 0, self)
        self.progress.setStyleSheet(self.STYLESHEET_TEMPLATE)
        self.progress.canceled.connect(self.cancel_task)
        self.progress.setWindowModality(Qt.WindowModal)
        
        self.thread = QThread(); self.worker = EmailWorker(task, smtp_details); self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.process)
        self.worker.finished.connect(self.handle_task_finished)
        self.worker.finished.connect(self.thread.quit)

        # Cleanup is tied to the thread's lifecycle to prevent race conditions
        self.thread.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.finished.connect(self.cleanup_thread)

        self.worker.progress.connect(self.update_email_progress)
        self.send_emails_button.setEnabled(False)
        self.thread.start(); self.progress.show()

    def update_email_progress(self, current, total):
        if self.progress.maximum() != total:
            self.progress.setMaximum(total)
        self.progress.setValue(current)

    def handle_task_finished(self, status, message):
        """Handles success, error, and cancellation feedback."""
        if self.progress:
            self.progress.close()
            self.progress = None # Explicitly clear the reference

        # IMPROVED FEEDBACK: Provide explicit confirmation for cancellation.
        is_cancellation = "cancel" in message.lower()
        if is_cancellation:
            QMessageBox.information(self, "Task Cancelled", "The task was successfully stopped.")
        elif status == 'success':
            QMessageBox.information(self, "Task Complete", message)
        else: # This handles actual errors
            self.show_error(message)

    def cleanup_thread(self):
        """Resets the thread and worker attributes and re-enables action buttons."""
        self.thread = None
        self.worker = None
        if hasattr(self, 'send_emails_button'):
            self.send_emails_button.setEnabled(True)
        if hasattr(self, 'download_drive_button'):
            self.download_drive_button.setEnabled(True)
        if hasattr(self, 'generate_tracker_button'):
            self.generate_tracker_button.setEnabled(True)
        if hasattr(self, 'update_form_button'):
            self.update_form_button.setEnabled(True)

    def setup_drive_tab(self):
        tab = QWidget(); layout = QVBoxLayout(tab); grid = QGridLayout()
        grid.setSpacing(10)
        grid.addWidget(QLabel("Folder Task Title:"), 0, 0); self.drive_task_title = QLineEdit(); grid.addWidget(self.drive_task_title, 0, 1, 1, 2)
        grid.addWidget(QLabel("Google Drive Folder URL or ID:"), 1, 0); self.drive_folder_id = QLineEdit(); grid.addWidget(self.drive_folder_id, 1, 1, 1, 2)
        grid.addWidget(QLabel("Download To (Folder):"), 2, 0); self.drive_local_path = QLineEdit(); grid.addWidget(self.drive_local_path, 2, 1)
        browse = QPushButton("Browse"); browse.clicked.connect(lambda: self.drive_local_path.setText(QFileDialog.getExistingDirectory(self, "Choose Folder"))); grid.addWidget(browse, 2, 2)
        grid.setColumnStretch(1, 1)
        layout.addLayout(grid)
        buttons_layout = QHBoxLayout(); save_btn = QPushButton("Save/Update Folder Task"); save_btn.clicked.connect(self.save_drive_task); delete_btn = QPushButton("Delete Folder Task"); delete_btn.setObjectName("DeleteButton"); delete_btn.clicked.connect(self.delete_drive_task); buttons_layout.addWidget(save_btn); buttons_layout.addWidget(delete_btn); layout.addLayout(buttons_layout)
        layout.addWidget(QLabel("Existing Folder Tasks:")); self.drive_task_combo = QComboBox(); self.drive_task_combo.currentIndexChanged.connect(self.load_selected_drive_task); layout.addWidget(self.drive_task_combo)
        layout.addStretch()
        self.download_drive_button = QPushButton("📥 Download Folder Now"); self.download_drive_button.setObjectName("PrimaryButton"); self.download_drive_button.clicked.connect(self.download_drive_click); layout.addWidget(self.download_drive_button)
        self.tabs.addTab(tab, "📁 Drive Folder")

    def setup_tracker_tab(self):
        tab = QWidget(); layout = QVBoxLayout(tab); grid = QGridLayout()
        grid.setSpacing(10)
        grid.addWidget(QLabel("Tracking Task Title:"), 0, 0); self.track_title = QLineEdit(); grid.addWidget(self.track_title, 0, 1, 1, 2)
        grid.addWidget(QLabel("Master Excel:"), 1, 0); self.master_excel_path = QLineEdit(); grid.addWidget(self.master_excel_path, 1, 1)
        master_browse = QPushButton("Browse"); master_browse.clicked.connect(lambda: self.master_excel_path.setText(QFileDialog.getOpenFileName(self, "Select Master Excel", "", "Excel Files (*.xlsx *.xls)")[0])); grid.addWidget(master_browse, 1, 2)
        grid.addWidget(QLabel("Response Sheet URL or ID:"), 2, 0); self.response_sheet_id = QLineEdit(); grid.addWidget(self.response_sheet_id, 2, 1, 1, 2)
        grid.addWidget(QLabel("Save Tracker Excel File:"), 3, 0); self.tracker_output_path = QLineEdit(); grid.addWidget(self.tracker_output_path, 3, 1)
        out_browse = QPushButton("Save As"); out_browse.clicked.connect(lambda: self.tracker_output_path.setText(QFileDialog.getSaveFileName(self, "Save Tracker Excel", "", "Excel Files (*.xlsx)")[0])); grid.addWidget(out_browse, 3, 2)
        grid.setColumnStretch(1, 1)
        layout.addLayout(grid)
        buttons_layout = QHBoxLayout(); save_btn = QPushButton("Save/Update Tracker Task"); save_btn.clicked.connect(self.save_tracker_task); delete_btn = QPushButton("Delete Tracker Task"); delete_btn.setObjectName("DeleteButton"); delete_btn.clicked.connect(self.delete_tracker_task); buttons_layout.addWidget(save_btn); buttons_layout.addWidget(delete_btn); layout.addLayout(buttons_layout)
        layout.addWidget(QLabel("Existing Tracker Tasks:"))
        self.track_select_combo = QComboBox(); self.track_select_combo.currentIndexChanged.connect(self.load_selected_tracker); layout.addWidget(self.track_select_combo)
        layout.addStretch()
        self.generate_tracker_button = QPushButton("📊 Generate Tracker Now"); self.generate_tracker_button.setObjectName("PrimaryButton"); self.generate_tracker_button.clicked.connect(self.generate_tracker_click); layout.addWidget(self.generate_tracker_button)
        self.tabs.addTab(tab, "📈 Tracker")

    def setup_form_updater_tab(self):
        tab = QWidget(); layout = QVBoxLayout(tab); grid = QGridLayout()
        grid.setSpacing(10)
        grid.addWidget(QLabel("Task Title:"), 0, 0); self.form_task_title = QLineEdit(); grid.addWidget(self.form_task_title, 0, 1)
        grid.addWidget(QLabel("Select Tracker Task (for Data Source):"), 1, 0); self.form_tracker_combo = QComboBox(); self.form_tracker_combo.currentIndexChanged.connect(self.load_form_tracker_path); grid.addWidget(self.form_tracker_combo, 1, 1)
        grid.addWidget(QLabel("Master Excel Path:"), 2, 0); self.form_excel_path = QLineEdit(); self.form_excel_path.setReadOnly(True); grid.addWidget(self.form_excel_path, 2, 1)
        grid.addWidget(QLabel("Google Form Link:"), 3, 0); self.form_link_input = QLineEdit(); self.form_link_input.setPlaceholderText("https://docs.google.com/forms/d/e/FORM_ID/viewform?usp=sf_link"); grid.addWidget(self.form_link_input, 3, 1)
        grid.setColumnStretch(1, 1)
        layout.addLayout(grid)
        buttons_layout = QHBoxLayout(); save_btn = QPushButton("Save/Update Form Task"); save_btn.clicked.connect(self.save_form_task); delete_btn = QPushButton("Delete Form Task"); delete_btn.setObjectName("DeleteButton"); delete_btn.clicked.connect(self.delete_form_task); buttons_layout.addWidget(save_btn); buttons_layout.addWidget(delete_btn); layout.addLayout(buttons_layout)
        layout.addWidget(QLabel("Existing Form Updater Tasks:"))
        self.form_task_combo = QComboBox(); self.form_task_combo.currentIndexChanged.connect(self.load_selected_form_task); layout.addWidget(self.form_task_combo)
        note_label = QLabel("<b>Note:</b> This requires editor access to the Google Form."); note_label.setWordWrap(True); layout.addWidget(note_label)
        layout.addStretch()
        self.update_form_button = QPushButton("🚀 Update Form Dropdowns Now"); self.update_form_button.setObjectName("PrimaryButton"); self.update_form_button.clicked.connect(self.update_form_click); layout.addWidget(self.update_form_button)
        self.tabs.addTab(tab, "📝 Form Updater")

    def setup_reminder_tab(self):
        # This function remains unchanged
        tab = QWidget(); layout = QVBoxLayout(tab); grid = QGridLayout()
        grid.setSpacing(10)
        grid.addWidget(QLabel("Reminder Title:"), 0, 0); self.reminder_title = QLineEdit(); grid.addWidget(self.reminder_title, 0, 1, 1, 3)
        grid.addWidget(QLabel("Subject:"), 1, 0); self.reminder_subject = QLineEdit(); grid.addWidget(self.reminder_subject, 1, 1, 1, 3)
        grid.addWidget(QLabel("Select Tracker (for Emails):"), 2, 0); self.reminder_track_combo = QComboBox(); self.reminder_track_combo.currentIndexChanged.connect(self.fill_reminder_track_path); grid.addWidget(self.reminder_track_combo, 2, 1, 1, 3)
        grid.addWidget(QLabel("Tracker Excel Path:"), 3, 0); self.reminder_excel_path = QLineEdit(); self.reminder_excel_path.setReadOnly(True); grid.addWidget(self.reminder_excel_path, 3, 1, 1, 3)
        grid.addWidget(QLabel("Start Date:"), 4, 0); self.reminder_start_date = QDateEdit(); self.reminder_start_date.setCalendarPopup(True); self.reminder_start_date.setDisplayFormat("yyyy-MM-dd"); self.reminder_start_date.setDate(datetime.datetime.today()); grid.addWidget(self.reminder_start_date, 4, 1)
        grid.addWidget(QLabel("End Date:"), 4, 2); self.reminder_end_date = QDateEdit(); self.reminder_end_date.setCalendarPopup(True); self.reminder_end_date.setDisplayFormat("yyyy-MM-dd"); self.reminder_end_date.setDate(datetime.datetime.today()); grid.addWidget(self.reminder_end_date, 4, 3)
        grid.addWidget(QLabel("Send Email On:"), 5, 0); self.reminder_freq_combo = QComboBox(); self.reminder_freq_combo.addItems(["Everyday", "Select Dates"]); grid.addWidget(self.reminder_freq_combo, 5, 1)
        grid.addWidget(QLabel("If select, enter dates (csv):"), 5, 2); self.reminder_dates = QLineEdit(); grid.addWidget(self.reminder_dates, 5, 3)
        grid.addWidget(QLabel("Send CC After (Days):"), 6, 2); self.reminder_cc_days = QLineEdit(); grid.addWidget(self.reminder_cc_days, 6, 3)
        grid.addWidget(QLabel("Message to Send:"), 7, 0, Qt.AlignTop); self.reminder_msg = QTextEdit(); self.reminder_msg.setMinimumHeight(80); grid.addWidget(self.reminder_msg, 7, 1, 1, 3)
        grid.setColumnStretch(1, 1); grid.setColumnStretch(3, 1)
        layout.addLayout(grid)
        buttons_layout = QHBoxLayout(); save_btn = QPushButton("Save/Update Reminder"); save_btn.clicked.connect(self.save_reminder); delete_btn = QPushButton("Delete Reminder"); delete_btn.setObjectName("DeleteButton"); delete_btn.clicked.connect(self.delete_reminder); buttons_layout.addWidget(save_btn); buttons_layout.addWidget(delete_btn); layout.addLayout(buttons_layout)
        layout.addWidget(QLabel("Existing Reminders:"))
        self.edit_reminder_combo = QComboBox(); self.edit_reminder_combo.currentIndexChanged.connect(self.load_selected_reminder); layout.addWidget(self.edit_reminder_combo)
        note = QLabel("<b>Note:</b> CC emails are sent from the 'CC Email ID' column in the Master Excel file of the selected Tracker Task."); note.setWordWrap(True); layout.addWidget(note)
        layout.addStretch()
        self.tabs.addTab(tab, "🔔 Reminder")

    def reload_drive_combo(self):
        current_text = self.drive_task_combo.currentText()
        self.drive_task_combo.blockSignals(True)
        self.drive_task_combo.clear(); self.config.setdefault("drive_tasks", []); self.drive_task_combo.addItem("--- Create New Task ---"); self.drive_task_combo.addItems([t.get("title") for t in self.config["drive_tasks"]])
        index = self.drive_task_combo.findText(current_text)
        if index != -1: self.drive_task_combo.setCurrentIndex(index)
        self.drive_task_combo.blockSignals(False)
    def save_drive_task(self):
        title, folder_id, path = self.drive_task_title.text().strip(), self.drive_folder_id.text().strip(), self.drive_local_path.text().strip()
        if not all([title, folder_id, path]): self.show_error("All fields are required."); return
        task = {"title": title, "folder_id": folder_id, "path": path, "type": "drive"}
        tasks = self.config.setdefault("drive_tasks", []); self.config["drive_tasks"] = [t for t in tasks if t.get("title") != title]; self.config["drive_tasks"].append(task)
        self.save_and_reload(); QMessageBox.information(self, "Saved", "Drive folder task saved.")
        self.drive_task_combo.setCurrentText(title)
    def delete_drive_task(self):
        title = self.drive_task_combo.currentText()
        if not title or title == "--- Create New Task ---": self.show_error("Select a task to delete"); return
        reply = QMessageBox.question(self, 'Confirm Delete', f"Are you sure you want to delete the task '{title}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No: return
        tasks = self.config.get("drive_tasks", []); self.config["drive_tasks"] = [t for t in tasks if t.get("title") != title]
        self.save_and_reload(); 
        self.load_selected_drive_task()
        QMessageBox.information(self, "Deleted", "Drive folder task deleted.")
    def load_selected_drive_task(self):
        title = self.drive_task_combo.currentText()
        obj = next((t for t in self.config.get("drive_tasks", []) if t.get("title") == title), None)
        if obj: self.drive_task_title.setText(obj.get("title", "")); self.drive_folder_id.setText(obj.get("folder_id", "")); self.drive_local_path.setText(obj.get("path", ""))
        else: self.drive_task_title.clear(); self.drive_folder_id.clear(); self.drive_local_path.clear()
    def download_drive_click(self):
        if self.thread and self.thread.isRunning(): self.show_error("A task is already running."); return
        user_input = self.drive_folder_id.text().strip()
        folder_id = get_google_id_from_url(user_input)
        path = self.drive_local_path.text().strip()
        if not folder_id or not path: self.show_error("Folder ID/URL and a local path are required."); return
        creds = get_creds()
        if not creds: return
        self.progress = QProgressDialog("Downloading folder...", "Cancel", 0, 100, self); self.progress.setStyleSheet(self.STYLESHEET_TEMPLATE); self.progress.canceled.connect(self.cancel_task); self.progress.setWindowModality(Qt.WindowModal)
        self.thread = QThread(); self.worker = DriveWorker(creds, folder_id, path); self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.process)
        self.worker.finished.connect(self.handle_task_finished)
        self.worker.finished.connect(self.thread.quit)

        # Cleanup is tied to the thread's lifecycle to prevent race conditions
        self.thread.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.finished.connect(self.cleanup_thread)

        self.worker.progress.connect(self.update_drive_progress)
        self.download_drive_button.setEnabled(False)
        self.thread.start(); self.progress.show()
    def update_drive_progress(self, current, total): self.progress.setMaximum(total); self.progress.setValue(current)
    
    def reload_tracker_combo(self):
        current_text = self.track_select_combo.currentText()
        self.track_select_combo.blockSignals(True)
        self.track_select_combo.clear(); self.config.setdefault("track_tasks", []); self.track_select_combo.addItem("--- Create New Task ---"); self.track_select_combo.addItems([t.get("title","") for t in self.config["track_tasks"]])
        index = self.track_select_combo.findText(current_text)
        if index != -1: self.track_select_combo.setCurrentIndex(index)
        self.track_select_combo.blockSignals(False)
    def save_tracker_task(self):
        title, master_path, response_sheet, out_path = self.track_title.text().strip(), self.master_excel_path.text().strip(), self.response_sheet_id.text().strip(), self.tracker_output_path.text().strip()
        if not all([title, master_path, response_sheet, out_path]): self.show_error("Title, master excel, response sheet, and tracker save path are required."); return
        task = { "title": title, "master_excel": master_path, "response_sheet_id": response_sheet, "result_path": out_path, "type": "tracker" }
        tasks = self.config.setdefault("track_tasks", []); self.config["track_tasks"] = [t for t in tasks if t.get("title") != title]; self.config["track_tasks"].append(task)
        self.save_and_reload(); QMessageBox.information(self, "Saved", "Tracker task saved.")
        self.track_select_combo.setCurrentText(title)
        
    def delete_tracker_task(self):
        """Deletes a tracker task and any linked reminder or form updater tasks."""
        title = self.track_select_combo.currentText()
        if not title or title == "--- Create New Task ---":
            self.show_error("Select a tracker task to delete.")
            return

        reply = QMessageBox.question(self, 'Confirm Delete',
                                     f"Are you sure you want to delete the tracker '{title}'?\n\n"
                                     f"This will also delete any Reminder or Form Updater tasks linked to it.",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            return

        tasks = self.config.get("track_tasks", [])
        self.config["track_tasks"] = [t for t in tasks if t.get("title") != title]

        deleted_reminders = 0
        reminders = self.config.get("reminders", [])
        original_reminder_count = len(reminders)
        self.config["reminders"] = [r for r in reminders if r.get("tracker_title") != title]
        deleted_reminders = original_reminder_count - len(self.config["reminders"])

        deleted_form_updaters = 0
        form_updaters = self.config.get("form_updater_tasks", [])
        original_form_updater_count = len(form_updaters)
        self.config["form_updater_tasks"] = [f for f in form_updaters if f.get("tracker_title") != title]
        deleted_form_updaters = original_form_updater_count - len(self.config["form_updater_tasks"])

        self.save_and_reload()
        self.load_selected_tracker()

        info_message = f"Tracker task '{title}' deleted."
        if deleted_reminders > 0:
            info_message += f"\nAlso removed {deleted_reminders} linked Reminder task(s)."
        if deleted_form_updaters > 0:
            info_message += f"\nAlso removed {deleted_form_updaters} linked Form Updater task(s)."
        QMessageBox.information(self, "Deleted", info_message)

    def load_selected_tracker(self):
        title = self.track_select_combo.currentText()
        obj = next((t for t in self.config.get("track_tasks", []) if t.get("title") == title), None)
        if obj:
            self.track_title.setText(obj.get("title", "")); self.master_excel_path.setText(obj.get("master_excel", "")); self.response_sheet_id.setText(obj.get("response_sheet_id", "")); self.tracker_output_path.setText(obj.get("result_path", ""))
        else:
            self.track_title.clear(); self.master_excel_path.clear(); self.response_sheet_id.clear(); self.tracker_output_path.clear()
    def generate_tracker_click(self):
        if self.thread and self.thread.isRunning(): self.show_error("A task is already running."); return
        response_input = self.response_sheet_id.text().strip()
        sheet_id = get_google_id_from_url(response_input)
        task_details = {"master_excel": self.master_excel_path.text().strip(), "response_sheet_id": sheet_id, "result_path": self.tracker_output_path.text().strip()}
        if not all(task_details.values()): self.show_error("All fields are required."); return
        creds = get_creds()
        if not creds: return
        self.progress = QProgressDialog("Generating tracker file...", "Cancel", 0, 0, self); self.progress.setStyleSheet(self.STYLESHEET_TEMPLATE); self.progress.canceled.connect(self.cancel_task); self.progress.setWindowModality(Qt.WindowModal)
        self.thread = QThread(); self.worker = TrackerWorker(creds, task_details); self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.process)
        self.worker.finished.connect(self.handle_task_finished)
        self.worker.finished.connect(self.thread.quit)

        # Cleanup is tied to the thread's lifecycle to prevent race conditions
        self.thread.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.finished.connect(self.cleanup_thread)

        self.generate_tracker_button.setEnabled(False)
        self.thread.start(); self.progress.show()

    def reload_form_updater_combo(self):
        current_text = self.form_task_combo.currentText()
        self.form_task_combo.blockSignals(True)
        self.form_task_combo.clear(); self.config.setdefault("form_updater_tasks", []); self.form_task_combo.addItem("--- Create New Task ---"); self.form_task_combo.addItems([t.get("title","") for t in self.config["form_updater_tasks"]])
        index = self.form_task_combo.findText(current_text)
        if index != -1: self.form_task_combo.setCurrentIndex(index)
        self.form_task_combo.blockSignals(False)
    def save_form_task(self):
        title, tracker_title, form_link = self.form_task_title.text().strip(), self.form_tracker_combo.currentText(), self.form_link_input.text().strip()
        if not all([title, tracker_title, form_link]): self.show_error("Task Title, a selected Tracker Task, and a Form Link are required."); return
        task = {"title": title, "tracker_title": tracker_title, "form_link": form_link, "type": "form_updater"}
        tasks = self.config.setdefault("form_updater_tasks", []); self.config["form_updater_tasks"] = [t for t in tasks if t.get("title") != title]; self.config["form_updater_tasks"].append(task)
        self.save_and_reload(); QMessageBox.information(self, "Saved", "Form Updater task saved.")
        self.form_task_combo.setCurrentText(title)
    def delete_form_task(self):
        title = self.form_task_combo.currentText()
        if not title or title == "--- Create New Task ---": self.show_error("Select a task to delete."); return
        reply = QMessageBox.question(self, 'Confirm Delete', f"Are you sure you want to delete the task '{title}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No: return
        tasks = self.config.get("form_updater_tasks", []); self.config["form_updater_tasks"] = [t for t in tasks if t.get("title") != title]
        self.save_and_reload();
        self.load_selected_form_task()
        QMessageBox.information(self, "Deleted", "Form Updater task deleted.")
    def load_selected_form_task(self):
        title = self.form_task_combo.currentText()
        task = next((t for t in self.config.get("form_updater_tasks", []) if t.get("title") == title), None)
        if task:
            self.form_task_title.setText(task.get("title", "")); self.form_tracker_combo.setCurrentText(task.get("tracker_title", "")); self.form_link_input.setText(task.get("form_link", ""))
        else:
            self.form_task_title.clear(); self.form_tracker_combo.setCurrentIndex(0); self.form_link_input.clear()
    def load_form_tracker_path(self):
        title = self.form_tracker_combo.currentText()
        task = next((t for t in self.config.get("track_tasks", []) if t.get("title") == title), None)
        self.form_excel_path.setText(task.get("master_excel", "") if task else "")
    def reload_form_tracker_combo(self):
        current_text = self.form_tracker_combo.currentText()
        self.form_tracker_combo.blockSignals(True)
        self.form_tracker_combo.clear(); self.config.setdefault("track_tasks", []); self.form_tracker_combo.addItem(""); self.form_tracker_combo.addItems([t.get("title","") for t in self.config["track_tasks"]])
        index = self.form_tracker_combo.findText(current_text)
        if index != -1: self.form_tracker_combo.setCurrentIndex(index)
        self.form_tracker_combo.blockSignals(False)
    def update_form_click(self):
        if self.thread and self.thread.isRunning(): self.show_error("A task is already running."); return
        excel_path, form_link = self.form_excel_path.text().strip(), self.form_link_input.text().strip()
        if not excel_path or not form_link: self.show_error("Please select a tracker task and provide a Google Form link."); return
        form_id = get_google_id_from_url(form_link)
        creds = get_creds()
        if not creds: return
        self.progress = QProgressDialog("Updating Google Form...", "Cancel", 0, 0, self); self.progress.setStyleSheet(self.STYLESHEET_TEMPLATE); self.progress.canceled.connect(self.cancel_task); self.progress.setWindowModality(Qt.WindowModal)
        self.thread = QThread(); self.worker = FormWorker(creds, form_id, excel_path); self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.process)
        self.worker.finished.connect(self.handle_task_finished)
        self.worker.finished.connect(self.thread.quit)

        # Cleanup is tied to the thread's lifecycle to prevent race conditions
        self.thread.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.finished.connect(self.cleanup_thread)

        self.update_form_button.setEnabled(False)
        self.thread.start(); self.progress.show()

    def reload_reminder_tracker_combo(self):
        current_text = self.reminder_track_combo.currentText()
        self.reminder_track_combo.blockSignals(True)
        self.reminder_track_combo.clear(); self.config.setdefault("track_tasks", []); self.reminder_track_combo.addItem(""); self.reminder_track_combo.addItems([t.get("title","") for t in self.config["track_tasks"]])
        index = self.reminder_track_combo.findText(current_text)
        if index != -1: self.reminder_track_combo.setCurrentIndex(index)
        self.reminder_track_combo.blockSignals(False)
    def save_reminder(self):
        title, subject, msg = self.reminder_title.text().strip(), self.reminder_subject.text().strip(), self.reminder_msg.toPlainText().strip()
        cc_days = self.reminder_cc_days.text().strip()
        if not all([title, subject, msg]): self.show_error("Title, Subject, and Message are required."); return
        try: cc_days_val = int(cc_days) if cc_days else 0
        except ValueError: self.show_error("Please enter a valid number for 'Send CC After (Days)'."); return
        reminder = { "title": title, "subject": subject, "tracker_title": self.reminder_track_combo.currentText().strip(), "start_date": self.reminder_start_date.date().toString("yyyy-MM-dd"), "end_date": self.reminder_end_date.date().toString("yyyy-MM-dd"), "frequency": self.reminder_freq_combo.currentText(), "dates": [d.strip() for d in self.reminder_dates.text().replace(',', ' ').split() if d.strip()], "cc_days": cc_days_val, "message": msg,"type": "reminder"}
        tasks = self.config.setdefault("reminders", []); self.config["reminders"] = [r for r in tasks if r.get("title") != title]; self.config["reminders"].append(reminder)
        self.save_and_reload(); QMessageBox.information(self, "Saved", "Reminder saved.")
        self.edit_reminder_combo.setCurrentText(title)
    def delete_reminder(self):
        title = self.edit_reminder_combo.currentText()
        if not title or title == "--- Create New Task ---": self.show_error("Select a task to delete."); return
        reply = QMessageBox.question(self, 'Confirm Delete', f"Are you sure you want to delete the task '{title}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No: return
        tasks = self.config.get("reminders", []); self.config["reminders"] = [r for r in tasks if r.get("title") != title]
        self.save_and_reload(); 
        self.load_selected_reminder()
        QMessageBox.information(self, "Deleted", "Reminder deleted.")
    def load_selected_reminder(self):
        title = self.edit_reminder_combo.currentText()
        obj = next((r for r in self.config.get("reminders", []) if r.get("title") == title), None)
        if obj:
            self.reminder_title.setText(obj.get("title", "")); self.reminder_subject.setText(obj.get("subject", "")); self.reminder_track_combo.setCurrentText(obj.get("tracker_title", ""))
            self.reminder_start_date.setDate(datetime.datetime.strptime(obj.get("start_date"), "%Y-%m-%d").date()); self.reminder_end_date.setDate(datetime.datetime.strptime(obj.get("end_date"), "%Y-%m-%d").date())
            self.reminder_freq_combo.setCurrentText(obj.get("frequency", "Everyday")); self.reminder_dates.setText(" ".join(obj.get("dates", []))); self.reminder_cc_days.setText(str(obj.get("cc_days", ""))); self.reminder_msg.setPlainText(obj.get("message", ""))
        else:
            for w in [self.reminder_title, self.reminder_subject, self.reminder_excel_path, self.reminder_dates, self.reminder_cc_days, self.reminder_msg]: w.clear()
            self.reminder_track_combo.setCurrentIndex(0); self.reminder_start_date.setDate(datetime.datetime.today()); self.reminder_end_date.setDate(datetime.datetime.today()); self.reminder_freq_combo.setCurrentIndex(0)
    def fill_reminder_track_path(self):
        title = self.reminder_track_combo.currentText()
        task = next((t for t in self.config.get("track_tasks", []) if t.get("title") == title), None)
        self.reminder_excel_path.setText(task.get("master_excel", "") if task else "")
    def reload_reminder_combo(self):
        current_text = self.edit_reminder_combo.currentText()
        self.edit_reminder_combo.blockSignals(True)
        self.edit_reminder_combo.clear(); self.config.setdefault("reminders", []); self.edit_reminder_combo.addItem("--- Create New Task ---"); self.edit_reminder_combo.addItems([r.get("title","") for r in self.config["reminders"]])
        index = self.edit_reminder_combo.findText(current_text)
        if index != -1: self.edit_reminder_combo.setCurrentIndex(index)
        self.edit_reminder_combo.blockSignals(False)
    
    def refresh_dropdowns(self):
        self.reload_email_combo()
        self.reload_drive_combo()
        self.reload_tracker_combo()
        self.reload_reminder_tracker_combo()
        self.reload_reminder_combo()
        self.reload_form_tracker_combo()
        self.reload_form_updater_combo()
        self.refresh_dashboard_display()

if __name__ == "__main__":
    # Download the default background on first run
    download_default_background()
    
    app = QApplication(sys.argv)
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    
    app_window = AutomationApp()
    # Show the window maximized instead of fullscreen
    app_window.showMaximized()
    sys.exit(app.exec_())
