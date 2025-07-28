import sys
import os
import json
import logging
import smtplib
import io
import pandas as pd
import datetime
import time
import hashlib
import re

# --- Third-party libraries ---
# Make sure to install them: pip install pandas openpyxl google-api-python-client google-auth-oauthlib google-auth-httplib2 python-dotenv
from email.mime.text import MIMEText
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from googleapiclient.discovery import build
from googleapiclient import errors
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.http import MediaIoBaseDownload
from dotenv import load_dotenv

# --- Configuration ---
# This section should mirror the GUI's configuration
load_dotenv()

CONFIG_FILE = 'task_log.json'
STATE_FILE = 'headless_state.json' # Stores last run times and states
TOKEN_FILE = 'token.json'
CREDENTIALS_FILE = 'credentials.json'
LOGFILE = 'automation.log'

SMTP_EMAIL = os.environ.get('AUTOMATION_SMTP_EMAIL')
SMTP_PASSWORD = os.environ.get('AUTOMATION_SMTP_PASSWORD')
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SCOPES = ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/forms.body']

# Setup logging to append to the same log file as the GUI
# Add a handler to also print logs to the console for immediate feedback
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - HEADLESS - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOGFILE),
        logging.StreamHandler(sys.stdout)
    ]
)

# --- Configuration & State Management ---
def load_config():
    """Loads the main configuration file (task_log.json)."""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        except json.JSONDecodeError:
            logging.error(f"Error decoding JSON from {CONFIG_FILE}. The file might be corrupt.")
            return {}
    logging.warning(f"Configuration file {CONFIG_FILE} not found.")
    return {}

def load_state():
    """Loads the state of previously run tasks."""
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r') as f:
                return json.load(f)
        except json.JSONDecodeError:
            return {} # Return empty state if file is corrupt
    return {}

def save_state(state):
    """Saves the current state of tasks."""
    with open(STATE_FILE, 'w') as f:
        json.dump(state, f, indent=4)

# --- Core Logic Functions ---
def get_google_id_from_url(url):
    """Extracts the ID from a Google Drive/Sheets/Forms URL using regex."""
    patterns = [
        r'/d/([a-zA-Z0-9-_]+)',
        r'/folders/([a-zA-Z0-9-_]+)',
        r'id=([a-zA-Z0-9-_]+)'
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return url

def get_creds():
    creds = None
    if not os.path.exists(CREDENTIALS_FILE):
        logging.critical(f"Credentials file ('{CREDENTIALS_FILE}') not found. Headless script cannot run.")
        return None
    if os.path.exists(TOKEN_FILE):
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        except Exception as e:
            logging.warning(f"Failed to load token.json: {e}. Will try to re-authenticate.")
            creds = None
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                logging.info("Google credentials refreshed successfully.")
            except Exception as e:
                logging.error(f"Could not refresh token. You may need to re-authenticate via the GUI. Error: {e}")
                return None
        else:
            logging.error("Credentials are not valid and cannot be refreshed. Please run the GUI to re-authenticate.")
            return None
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())
    return creds

def send_email(recipient, subject, body, config, cc_recipient=None):
    smtp_details = config.get('settings', {})
    from_email = smtp_details.get('smtp_email', SMTP_EMAIL)
    password = smtp_details.get('smtp_password', SMTP_PASSWORD)
    
    if not from_email or not password:
        logging.error("SMTP credentials not configured. Cannot send email.")
        return False
    try:
        msg = MIMEText(body)
        msg['Subject'] = subject
        msg['From'] = from_email
        msg['To'] = recipient
        
        all_recipients = [recipient]
        if cc_recipient:
            cc_list = [email.strip() for email in cc_recipient.split(',') if email.strip()]
            if cc_list:
                msg['Cc'] = ', '.join(cc_list)
                all_recipients.extend(cc_list)

        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(from_email, password)
            server.sendmail(from_email, all_recipients, msg.as_string())
        logging.info(f"Sent email to {recipient} (CC: {cc_recipient or 'None'}) with subject: {subject}")
        return True
    except Exception as e:
        logging.error(f"Failed to send email to {recipient}: {e}")
        return False

def get_file_hash(filepath):
    if not os.path.exists(filepath):
        return None
    sha256_hash = hashlib.sha256()
    with open(filepath, "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()

# --- Task Handlers ---

def handle_drive_tasks(creds, tasks, state):
    logging.info("Checking for Drive tasks...")
    drive_service = build('drive', 'v3', credentials=creds)
    
    for task in tasks:
        task_title = task.get('title')
        folder_url_or_id = task.get('folder_id')
        local_path = task.get('path')
        
        if not all([task_title, folder_url_or_id, local_path]):
            logging.warning(f"Skipping invalid Drive task: {task_title or 'Untitled'}")
            continue

        folder_id = get_google_id_from_url(folder_url_or_id)
        logging.info(f"Processing Drive task: '{task_title}' (Folder ID: {folder_id})")

        try:
            # Get metadata for all files in the folder
            results = drive_service.files().list(q=f"'{folder_id}' in parents and trashed=false", fields="files(id, name, mimeType, modifiedTime, md5Checksum)").execute()
            remote_files = results.get('files', [])
            if not os.path.exists(local_path): os.makedirs(local_path)
            
            task_state = state.setdefault('drive_tasks', {}).setdefault(task_title, {})

            for item in remote_files:
                file_name, file_id, mime_type, remote_mod_time, remote_md5 = item.get('name'), item.get('id'), item.get('mimeType'), item.get('modifiedTime'), item.get('md5Checksum')
                
                # Logic for handling Google Workspace file exports
                is_google_doc = mime_type.startswith('application/vnd.google-apps')
                export_map = {
                    'application/vnd.google-apps.document': {'mime': 'application/pdf', 'ext': '.pdf'},
                    'application/vnd.google-apps.spreadsheet': {'mime': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'ext': '.xlsx'},
                    'application/vnd.google-apps.presentation': {'mime': 'application/vnd.openxmlformats-officedocument.presentationml.presentation', 'ext': '.pptx'},
                }
                
                request = None
                local_file_path = os.path.join(local_path, file_name)
                
                if is_google_doc:
                    if mime_type in export_map:
                        export_details = export_map[mime_type]
                        base_name, _ = os.path.splitext(local_file_path)
                        local_file_path = base_name + export_details['ext']
                        
                        # For Google Docs, we check modification time as they don't have md5 checksums
                        if os.path.exists(local_file_path) and task_state.get(file_name, {}).get('modifiedTime') == remote_mod_time:
                            continue # Skip if modification time matches
                        
                        logging.info(f"Exporting/updating '{file_name}' as {export_details['ext']}...")
                        request = drive_service.files().export_media(fileId=file_id, mimeType=export_details['mime'])
                    else:
                        logging.warning(f"Skipping unsupported Google App file: {file_name} ({mime_type})")
                        continue
                else: # Standard file
                    if os.path.exists(local_file_path) and task_state.get(file_name, {}).get('md5') == remote_md5:
                        continue # Skip if MD5 hash matches
                    
                    logging.info(f"Downloading/updating '{file_name}'...")
                    request = drive_service.files().get_media(fileId=file_id)

                # Execute download/export if a request was created
                if request:
                    fh = io.BytesIO()
                    downloader = MediaIoBaseDownload(fh, request)
                    done = False
                    while not done: status, done = downloader.next_chunk()
                    with open(local_file_path, 'wb') as f: f.write(fh.getbuffer())
                    
                    # Update state with the correct metadata
                    task_state[file_name] = {'modifiedTime': remote_mod_time, 'md5': remote_md5}

        except errors.HttpError as e: logging.error(f"API Error for Drive task '{task_title}': {e}")
        except Exception as e: logging.error(f"Failed to process Drive task '{task_title}': {e}")

def handle_email_tasks(config, tasks, state):
    logging.info("Checking for scheduled Email tasks...")
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    
    for task in tasks:
        task_title, schedule_date = task.get('title'), task.get('date')
        if schedule_date != today_str: continue

        task_run_key = f"{task_title}_{today_str}"
        if state.get('email_tasks', {}).get(task_run_key):
            logging.info(f"Email task '{task_title}' for {today_str} has already been executed. Skipping.")
            continue
            
        logging.info(f"Executing scheduled email task: '{task_title}'")
        try:
            df = pd.read_excel(task["excel"])
            for _, row in df.iterrows():
                recipient = next((str(row[col]).strip() for col in ['Email', 'Email ID'] if col in row and pd.notna(row[col])), None)
                if not recipient: continue
                send_email(recipient, task['subject'], task['msg'], config, cc_recipient=task.get('cc'))
                time.sleep(1)

            state.setdefault('email_tasks', {})[task_run_key] = True
            logging.info(f"Finished email task '{task_title}'.")
        except Exception as e: logging.error(f"Failed to execute email task '{task_title}': {e}")

def handle_tracker_tasks(creds, tasks, state):
    logging.info("Checking for Tracker tasks...")
    sheets_service = build('sheets', 'v4', credentials=creds)
    for task in tasks:
        task_title, sheet_url_or_id, master_excel, result_path = task.get('title'), task.get('response_sheet_id'), task.get('master_excel'), task.get('result_path')
        if not all([task_title, sheet_url_or_id, master_excel, result_path]):
            logging.warning(f"Skipping invalid Tracker task: {task_title or 'Untitled'}")
            continue
        
        sheet_id = get_google_id_from_url(sheet_url_or_id)
        try:
            # Get master file hash
            master_hash = get_file_hash(master_excel)
            
            # Get response sheet data and create a hash from its content for reliable change detection
            spreadsheet_metadata = sheets_service.spreadsheets().get(spreadsheetId=sheet_id).execute()
            sheet_name = next((s['properties']['title'] for s in spreadsheet_metadata.get('sheets', []) if s['properties']['title'].startswith("Form Responses")), spreadsheet_metadata['sheets'][0]['properties']['title'])
            data = sheets_service.spreadsheets().values().get(spreadsheetId=sheet_id, range=sheet_name).execute().get('values', [])
            
            response_data_str = json.dumps(data)
            response_data_hash = hashlib.sha256(response_data_str.encode('utf-8')).hexdigest()

            task_state = state.setdefault('tracker_tasks', {}).setdefault(task_title, {})
            
            # RELIABLE CHECK: Compare hash of master file AND hash of response data content
            if task_state.get('last_master_hash') == master_hash and task_state.get('last_response_data_hash') == response_data_hash:
                logging.info(f"Tracker source for '{task_title}' has not changed. Skipping generation.")
                continue

            logging.info(f"Change detected for '{task_title}'. Regenerating tracker...")
            
            if not data or len(data) < 1: 
                logging.error(f"No data in response sheet for '{task_title}'."); 
                continue
            
            header = data[0]
            padded_rows = [row + [None] * (len(header) - len(row)) for row in data[1:]]
            df_response = pd.DataFrame(padded_rows, columns=header).loc[:,~pd.DataFrame(padded_rows, columns=header).columns.duplicated()]
            df_master = pd.read_excel(master_excel)
            
            df_response['Email'] = df_response['Email'].astype(str).str.strip().str.lower()
            df_response['Location'] = df_response['Location'].astype(str).str.strip().str.lower()
            ts_col = 'Timestamp' if 'Timestamp' in df_response.columns else df_response.columns[0]
            
            # UPDATED: This lambda now splits by comma OR newline to handle all multi-upload cases.
            agg_response = (df_response.groupby(['Email', 'Location']).agg(
                docs=('Upload the Applicable Documents', lambda x: [link.strip() for doc in x if pd.notna(doc) and str(doc).strip() for link in re.split(r'[,\n]', str(doc)) if link.strip()]),
                timestamp=(ts_col, 'max')
            ).reset_index())
            
            response_map = {(row['Email'], row['Location']): (row['docs'], row['timestamp']) for _, row in agg_response.iterrows()}
            
            final_rows, s_no = [], 1
            for _, row in df_master.iterrows():
                key = (str(row['Email ID']).strip().lower(), str(row['Location']).strip().lower())
                docs, timestamp = response_map.get(key, ([], ''))

                if docs:
                    first_row_in_group = True
                    for doc_link in docs:
                        final_rows.append({
                            'S.No.': s_no if first_row_in_group else None,
                            'Location': row['Location'] if first_row_in_group else None,
                            'SPOC': row['SPOC'] if first_row_in_group else None,
                            'Email ID': row['Email ID'] if first_row_in_group else None,
                            'Document Uploaded': doc_link,
                            'Uploaded': 'Yes' if first_row_in_group else None,
                            'Uploaded When': timestamp if first_row_in_group else None
                        })
                        first_row_in_group = False
                    s_no += 1
                else:
                    final_rows.append({
                        'S.No.': s_no, 'Location': row['Location'], 'SPOC': row['SPOC'], 
                        'Email ID': row['Email ID'], 'Document Uploaded': '', 'Uploaded': 'No', 'Uploaded When': ''
                    })
                    s_no += 1

            tracker_df = pd.DataFrame(final_rows, columns=['S.No.', 'Location', 'SPOC', 'Email ID', 'Document Uploaded', 'Uploaded', 'Uploaded When'])
            tracker_df.to_excel(result_path, index=False)

            # Apply advanced formatting with openpyxl
            wb = load_workbook(result_path)
            ws = wb.active
            
            for idx, width in enumerate([6, 20, 25, 30, 50, 12, 25], start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width
            
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            center_align = Alignment(vertical="center", horizontal="center", wrap_text=True)

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.column != 5: # Don't center the document links column
                        cell.alignment = center_align
            
            cols_to_merge = [1, 2, 3, 4, 6, 7]
            current_row_idx = 2
            while current_row_idx <= ws.max_row:
                if ws.cell(row=current_row_idx, column=1).value is not None:
                    merge_count = 1
                    scan_row_idx = current_row_idx + 1
                    while scan_row_idx <= ws.max_row and ws.cell(row=scan_row_idx, column=1).value is None:
                        merge_count += 1
                        scan_row_idx += 1
                    
                    uploaded_cell = ws.cell(row=current_row_idx, column=6)
                    if uploaded_cell.value == "Yes": uploaded_cell.fill = green_fill
                    elif uploaded_cell.value == "No": uploaded_cell.fill = red_fill

                    if merge_count > 1:
                        end_row_idx = current_row_idx + merge_count - 1
                        for col_idx in cols_to_merge:
                            ws.merge_cells(start_row=current_row_idx, start_column=col_idx, end_row=end_row_idx, end_column=col_idx)
                    
                    current_row_idx += merge_count
                else:
                    current_row_idx += 1
            
            wb.save(result_path)
            logging.info(f"Successfully generated tracker for '{task_title}' at {result_path}")
            
            # Save the new, reliable hashes to the state file
            task_state['last_master_hash'] = master_hash
            task_state['last_response_data_hash'] = response_data_hash
            task_state['last_generated'] = datetime.datetime.now().isoformat()

        except errors.HttpError as e: logging.error(f"API Error processing Tracker task '{task_title}': {e}")
        except Exception as e: logging.error(f"Failed to process Tracker task '{task_title}': {e}", exc_info=True)


def handle_form_updater_tasks(creds, tasks, state, all_configs):
    logging.info("Checking for Form Updater tasks...")
    forms_service = build('forms', 'v1', credentials=creds)
    for task in tasks:
        task_title, tracker_title, form_url_or_id = task.get('title'), task.get('tracker_title'), task.get('form_link')
        tracker_task = next((t for t in all_configs.get('track_tasks', []) if t.get('title') == tracker_title), None)
        if not all([task_title, form_url_or_id, tracker_task]):
            logging.warning(f"Skipping invalid Form Updater task '{task_title}'. Could not find linked tracker task '{tracker_title}'.")
            continue
        
        excel_path = tracker_task.get('master_excel')
        try:
            excel_hash = get_file_hash(excel_path)
            task_state = state.setdefault('form_updater_tasks', {}).setdefault(task_title, {})
            if task_state.get('last_excel_hash') == excel_hash:
                logging.info(f"Source Excel for Form Updater '{task_title}' has not changed. Skipping.")
                continue
            
            form_id = get_google_id_from_url(form_url_or_id)
            logging.info(f"Change detected for '{task_title}'. Updating form dropdowns... (Form ID: {form_id})")
            
            df = pd.read_excel(excel_path)
            field_mappings = {'Location': 'Location', 'Email': 'Email ID', 'SPOC Name': 'SPOC'}
            form = forms_service.forms().get(formId=form_id).execute()
            form_items = form.get('items', [])
            requests, updated_fields = [], []
            for form_title, excel_column in field_mappings.items():
                if excel_column not in df.columns: logging.warning(f"Column '{excel_column}' not in Excel. Skipping '{form_title}'."); continue
                options_list = df[excel_column].dropna().unique().tolist()
                if not options_list: logging.warning(f"No data in column '{excel_column}'. Skipping '{form_title}'."); continue
                target_item, target_index = None, -1
                for i, item in enumerate(form_items):
                    if item.get('title', '').strip().lower() == form_title.lower():
                        target_item, target_index = item, i; break
                if not target_item: logging.warning(f"Question '{form_title}' not in form. Skipping."); continue
                dropdown_options = [{'value': str(opt)} for opt in options_list]
                new_item_body = {
                    'itemId': target_item.get('itemId'), 'title': target_item.get('title'),
                    'questionItem': {'question': {
                        'questionId': target_item['questionItem']['question']['questionId'],
                        'required': True,
                        'choiceQuestion': {'type': 'DROP_DOWN', 'options': dropdown_options, 'shuffle': False}
                    }}}
                request = {'updateItem': {'item': new_item_body, 'location': {'index': target_index}, 'updateMask': 'questionItem'}}
                requests.append(request)
                updated_fields.append(form_title)
            if not requests: logging.warning(f"No matching questions or data to update in form for task '{task_title}'."); continue
            body = {'requests': requests}
            forms_service.forms().batchUpdate(formId=form_id, body=body).execute()
            logging.info(f"Successfully updated form for task '{task_title}'. Updated fields: {', '.join(updated_fields)}")
            
            task_state['last_excel_hash'] = excel_hash
            task_state['last_updated'] = datetime.datetime.now().isoformat()
        except Exception as e: logging.error(f"Failed to process Form Updater task '{task_title}': {e}")

def handle_reminder_tasks(config, tasks, state):
    logging.info("Checking for Reminder tasks...")
    today = datetime.date.today()
    today_str = today.strftime("%Y-%m-%d")
    
    for task in tasks:
        task_title = task.get('title')
        try:
            start_date = datetime.datetime.strptime(task['start_date'], "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(task['end_date'], "%Y-%m-%d").date()
            
            if not (start_date <= today <= end_date): continue

            is_send_day = False
            if task['frequency'] == 'Everyday':
                is_send_day = True
            elif task['frequency'] == 'Select Dates' and str(today.day) in task.get('dates', []):
                is_send_day = True
            
            if not is_send_day: continue

            logging.info(f"Processing reminders for task '{task_title}' on {today_str}")
            tracker_title = task.get('tracker_title')
            tracker_task = next((t for t in config.get('track_tasks', []) if t.get('title') == tracker_title), None)
            if not tracker_task or not os.path.exists(tracker_task.get('result_path')):
                logging.error(f"Could not find generated tracker file for reminder task '{task_title}'. Skipping.")
                continue

            df_tracker = pd.read_excel(tracker_task['result_path'])
            task_state = state.setdefault('reminder_tasks', {}).setdefault(task_title, {})
            
            for _, row in df_tracker.iterrows():
                if row['Uploaded'] == 'No':
                    email_id = row['Email ID']
                    if pd.isna(email_id): continue # Skip if email is blank
                    
                    if task_state.get(email_id) == today_str:
                        logging.info(f"Reminder for {email_id} already sent today. Skipping.")
                        continue
                    
                    logging.info(f"Sending reminder to {email_id} for task '{task_title}'.")
                    send_email(email_id, task['subject'], task['message'], config)
                    task_state[email_id] = today_str # Mark as sent for today
                    time.sleep(1)

        except Exception as e: logging.error(f"Failed to process Reminder task '{task_title}': {e}")

# --- Main Execution ---
def main():
    """Main function to run tasks in a continuous loop."""
    while True:
        logging.info("--- Starting new automation cycle ---")
        
        config = load_config()
        if not config:
            logging.warning("Configuration file is empty or not found. Waiting for next cycle.")
            time.sleep(60)
            continue
            
        state = load_state()
        creds = get_creds()
        
        if not creds:
            logging.critical("Could not obtain Google credentials. Waiting for next cycle.")
            time.sleep(60)
            continue

        # Process tasks if they exist in config
        if "drive_tasks" in config:
            handle_drive_tasks(creds, config["drive_tasks"], state)
        
        if "emails" in config:
            handle_email_tasks(config, config["emails"], state)
            
        if "track_tasks" in config:
            handle_tracker_tasks(creds, config["track_tasks"], state)

        if "form_updater_tasks" in config:
            handle_form_updater_tasks(creds, config["form_updater_tasks"], state, config)
            
        if "reminders" in config:
            handle_reminder_tasks(config, config["reminders"], state)

        # Save the updated state
        save_state(state)
        logging.info("--- Automation cycle finished. Waiting for 1 minute... ---")
        time.sleep(60)

if __name__ == "__main__":
    main()
